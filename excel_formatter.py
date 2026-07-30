# -*- coding: utf-8 -*-
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Border styles
thin_side = Side(style='thin', color='A0A0A0')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Colors matching user requirements & reference image
COLOR_RED = 'FF0000'         # Overdue (SLA breached)
COLOR_ORANGE = 'FFC000'      # Handover Today (Today SLA)
COLOR_DARK_RED = 'C00000'    # Order status at NEW
COLOR_GREEN = '92D050'       # Within SLA (Future)
COLOR_LIGHT_GREEN = 'E2EFDA'  # Not reflected in OMS
COLOR_GREY = 'D3D3D3'         # Unpaid Orders

FILL_RED = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type='solid')
FILL_ORANGE = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type='solid')
FILL_DARK_RED = PatternFill(start_color=COLOR_DARK_RED, end_color=COLOR_DARK_RED, fill_type='solid')
FILL_GREEN = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type='solid')
FILL_LIGHT_GREEN = PatternFill(start_color=COLOR_LIGHT_GREEN, end_color=COLOR_LIGHT_GREEN, fill_type='solid')
FILL_GREY = PatternFill(start_color=COLOR_GREY, end_color=COLOR_GREY, fill_type='solid')

FONT_WHITE_BOLD = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
FONT_BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')
FONT_NORMAL = Font(name='Calibri', size=11, bold=False)
FONT_BOLD = Font(name='Calibri', size=11, bold=True)
FONT_TITLE = Font(name='Calibri', size=14, bold=True)

def _is_blank(val):
    s = str(val).strip()
    return not s or s.lower() in ("nan", "none", "nat", "null")

def get_date_status(date_str, ref_date_str):
    """
    Compare date_str to ref_date_str (both in 'DD-MM-YYYY' format).
    Returns 'Breached', 'Today', 'Future', or 'Unknown'.
    """
    if _is_blank(date_str) or _is_blank(ref_date_str):
        return 'Unknown'
    try:
        dt = datetime.strptime(date_str.strip(), '%d-%m-%Y')
        ref_dt = datetime.strptime(ref_date_str.strip(), '%d-%m-%Y')
        if dt < ref_dt:
            return 'Breached'
        elif dt == ref_dt:
            return 'Today'
        else:
            return 'Future'
    except Exception:
        return 'Unknown'

def _ensure_puma_logo_exists():
    import os
    if not os.path.exists("puma_logo.png"):
        try:
            from PIL import Image as PILImage, ImageDraw
            img = PILImage.new('RGBA', (200, 80), color=(255, 255, 255, 255))
            draw = ImageDraw.Draw(img)
            
            # Arched body
            points = [
                (40, 45), (45, 40), (55, 30), (65, 25), (75, 22), (85, 22), 
                (95, 25), (105, 30), (120, 42), (135, 50), (150, 52), (160, 50),
                (150, 45), (135, 40), (120, 32), (105, 22), (95, 17), (80, 15),
                (65, 17), (55, 22), (45, 30), (35, 40), (25, 50), (30, 52)
            ]
            draw.polygon(points, fill=(16, 16, 16, 255))
            
            # Legs & tail
            draw.line([(55, 30), (35, 18)], fill=(16, 16, 16, 255), width=4)
            draw.line([(50, 33), (32, 23)], fill=(16, 16, 16, 255), width=3)
            draw.line([(135, 50), (155, 68)], fill=(16, 16, 16, 255), width=4)
            draw.line([(142, 51), (165, 64)], fill=(16, 16, 16, 255), width=3)
            draw.line([(150, 52), (180, 40), (190, 20)], fill=(16, 16, 16, 255), width=3)
            
            # Text PUMA
            draw.rectangle([60, 58, 64, 73], fill=(16, 16, 16, 255))
            draw.rectangle([60, 58, 72, 66], fill=(16, 16, 16, 255))
            draw.rectangle([70, 60, 72, 64], fill=(16, 16, 16, 255))
            
            draw.rectangle([78, 58, 82, 70], fill=(16, 16, 16, 255))
            draw.rectangle([88, 58, 92, 70], fill=(16, 16, 16, 255))
            draw.rectangle([78, 70, 92, 73], fill=(16, 16, 16, 255))
            
            draw.rectangle([98, 58, 102, 73], fill=(16, 16, 16, 255))
            draw.rectangle([112, 58, 116, 73], fill=(16, 16, 16, 255))
            draw.polygon([(102, 58), (107, 68), (112, 58), (109, 58), (107, 64), (105, 58)], fill=(16, 16, 16, 255))
            
            draw.rectangle([122, 60, 126, 73], fill=(16, 16, 16, 255))
            draw.rectangle([132, 60, 136, 73], fill=(16, 16, 16, 255))
            draw.rectangle([122, 58, 136, 61], fill=(16, 16, 16, 255))
            draw.rectangle([122, 66, 136, 68], fill=(16, 16, 16, 255))
            
            img.save("puma_logo.png")
        except Exception:
            pass

def autofit_columns(ws, df=None, min_width=10, padding=3):
    """Auto-adjusts columns width according to cells content length."""
    if df is not None and not df.empty:
        # Optimized vectorized logic from DataFrame for fast execution
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            lengths = df[col_name].astype(str).str.len()
            max_len = 0
            if not lengths.empty:
                val = lengths.max()
                if pd.notna(val):
                    max_len = int(val)
            header_len = len(str(col_name))
            ws.column_dimensions[col_letter].width = max(max(max_len, header_len) + padding, min_width)
    else:
        # Fallback cell-scan optimized by slicing first 100 rows for summary sheets
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col[:100]:
                val = str(cell.value or '')
                if '\n' in val:
                    lines_len = [len(l) for l in val.split('\n')]
                    cell_len = max(lines_len) if lines_len else 0
                else:
                    cell_len = len(val)
                if cell_len > max_len:
                    max_len = cell_len
                    
            ws.column_dimensions[col_letter].width = max(max_len + padding, min_width)

def format_data_sheet(ws, df):
    """Applies basic styling, bold headers, thin borders, and center alignment to data cells."""
    ws.sheet_view.showGridLines = True
    
    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    
    # Header Row formatting
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = FONT_BOLD
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
    # Data Rows formatting (optimized for large DataFrames)
    max_styled_rows = min(len(df) + 2, 500) if len(df) > 1000 else (len(df) + 2)
    for row_idx in range(2, max_styled_rows):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = FONT_NORMAL
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
                
    autofit_columns(ws, df)

def add_country_sheets_to_workbook(wb, country, raw_df, pivot_df, summary_df, ref_date_str):
    """
    Adds country styled Summary and Data sheets to an existing workbook.
    """
    ws_summary = wb.create_sheet(title=f"{country} Summary")
    ws_summary.sheet_view.showGridLines = True
    
    # ── Title Block merged to match the pivot table width ──
    title_end_col = len(pivot_df.columns) if not pivot_df.empty else 6
    ws_summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=title_end_col)
    ws_summary.row_dimensions[2].height = 24
    
    title_cell = ws_summary.cell(row=2, column=1, value=f"PUMA {country} - Pending Orders")
    
    # Custom PUMA brand styling (PUMA Red background, white bold text)
    puma_red_fill = PatternFill(start_color="BA0C2F", end_color="BA0C2F", fill_type="solid")
    font_title_white = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    
    title_cell.fill = puma_red_fill
    title_cell.font = font_title_white
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply background fill and borders around Title cell block
    for col_idx in range(1, title_end_col + 1):
        cell_block = ws_summary.cell(row=2, column=col_idx)
        cell_block.fill = puma_red_fill
        cell_block.border = thin_border
        
    # ── Pivot Table (Starting at A3 - no gap) ──
    if not pivot_df.empty:
        # Write headers
        headers = list(pivot_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_summary.cell(row=3, column=col_idx, value=header)
            cell.font = FONT_BOLD
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        # Write rows
        # Classic Pivot: consecutive duplicate channels are left blank
        last_channel = None
        for row_idx, row_data in enumerate(pivot_df.itertuples(index=False), start=4):
            is_grand_total_row = (row_data[0] == "Grand Total")
            
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
                # Check for Channel column (1st column)
                if col_idx == 1:
                    if is_grand_total_row:
                        cell.value = "Grand Total"
                        cell.font = FONT_BOLD
                    else:
                        chan_val = str(val or '').strip()
                        if chan_val == last_channel:
                            cell.value = ""
                        else:
                            cell.value = chan_val
                            last_channel = chan_val
                            cell.font = FONT_BOLD
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                # Check for OMS Status column (2nd column)
                elif col_idx == 2:
                    if is_grand_total_row:
                        cell.value = ""
                    else:
                        cell.value = str(val or '').strip()
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = FONT_BOLD if is_grand_total_row else FONT_NORMAL
                # Check for Date Columns & Grand Total value
                else:
                    cell.value = val
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Highlight counts > 0 based on date column SLA status
                    col_header = headers[col_idx - 1]
                    is_grand_total_col = (col_header == "Grand Total")
                    
                    if is_grand_total_row or is_grand_total_col:
                        cell.font = FONT_BOLD
                    else:
                        cell.font = FONT_NORMAL
                        # Highlight cell counts > 0 in data cells
                        if isinstance(val, (int, float)) and val > 0:
                            # Determine date status
                            status = get_date_status(col_header, ref_date_str)
                            if status == 'Today':
                                cell.fill = FILL_ORANGE
                            elif status == 'Future':
                                cell.fill = FILL_GREEN
                            elif status == 'Breached':
                                cell.fill = FILL_RED
                                cell.font = FONT_WHITE_BOLD
                                
    # ── Highlight Metrics Table (Starting at row 3 - no gap) ──
    start_col = (len(pivot_df.columns) + 2) if not pivot_df.empty else 11
    metrics_map = {}
    if not summary_df.empty:
        metrics_map = summary_df.set_index("Metric")["Count"].to_dict()
        
    metrics_list = [
        ("Overdue (SLA breached)", "Overdue", FILL_RED, FONT_WHITE_BOLD),
        ("Handover today (Today SLA)", "Handover Today", FILL_ORANGE, FONT_BLACK_BOLD),
        ("Order Status at New", "Order status at NEW", FILL_DARK_RED, FONT_WHITE_BOLD),
        ("Within SLA (Future)", "Within SLA", FILL_GREEN, FONT_BLACK_BOLD),
        ("Not reflecting in OM", "Not reflected in OMS", FILL_LIGHT_GREEN, FONT_BLACK_BOLD),
        ("Unpaid Orders", "Unpaid Orders", FILL_GREY, FONT_BLACK_BOLD)
    ]
    
    for idx, (original_name, display_name, fill, font) in enumerate(metrics_list):
        row_pos = 3 + idx
        
        # Metric Label
        cell_lbl = ws_summary.cell(row=row_pos, column=start_col, value=display_name)
        cell_lbl.fill = fill
        cell_lbl.font = font
        cell_lbl.alignment = Alignment(horizontal='center', vertical='center')
        cell_lbl.border = thin_border
        
        # Metric Value
        val = metrics_map.get(original_name, 0)
        if display_name == "Not reflected in OMS" and (val == 0 or val == "-"):
            val = "-"
            
        cell_val = ws_summary.cell(row=row_pos, column=start_col + 1, value=val)
        cell_val.font = FONT_BOLD
        cell_val.alignment = Alignment(horizontal='center', vertical='center')
        cell_val.border = thin_border
        
    autofit_columns(ws_summary)
    
    # == 2. Data Sheet =========================================================
    ws_data = wb.create_sheet(title=f"{country} Data")
    
    # Write Header
    if not raw_df.empty:
        cols = list(raw_df.columns)
        for col_idx, col_name in enumerate(cols, start=1):
            ws_data.cell(row=1, column=col_idx, value=col_name)
            
        # Write rows
        for row_idx, row_data in enumerate(raw_df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row_data, start=1):
                ws_data.cell(row=row_idx, column=col_idx, value=val)
                
        format_data_sheet(ws_data, raw_df)

def generate_excel_workbook(country, raw_df, pivot_df, summary_df, ref_date_str):
    """
    Creates a country-specific workbook (Summary + Data).
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    add_country_sheets_to_workbook(wb, country, raw_df, pivot_df, summary_df, ref_date_str)
    
    # Rename f"{country} Summary" -> "Summary" and f"{country} Data" -> "Data"
    wb.worksheets[0].title = "Summary"
    wb.worksheets[1].title = "Data"
    return wb

def generate_fast_excel_bytes(sheet_dict):
    """
    Ultra-fast Excel writer using xlsxwriter with openpyxl fallback.
    sheet_dict: dict of {"Sheet Name": dataframe}
    Returns bytes object of Excel file.
    """
    import io
    excel_buffer = io.BytesIO()
    
    try:
        with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
            workbook = writer.book
            header_fmt = workbook.add_format({
                'bold': True,
                'bg_color': '#87CEEB',  # Sky blue
                'font_color': '#000000',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            # Center alignment applied column-wide - harmless on blank cells
            # (no visible effect without a border or content), so this is
            # safe/fast to set via set_column.
            align_fmt = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter'
            })
            # Border-only format. NOTE: Excel's conditional-formatting (dxf)
            # spec does not support alignment, only border/font/fill/number
            # format - so border must be applied separately from alignment
            # via conditional_format if we want it scoped to just the data
            # rows instead of the entire column.
            border_fmt = workbook.add_format({'border': 1})
            
            for sheet_name, df in sheet_dict.items():
                if df is None:
                    df = pd.DataFrame()
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                
                # Format header row - sky blue, bold, centered, bordered
                for col_num, value in enumerate(df.columns):
                    worksheet.write(0, col_num, value, header_fmt)
                    
                # Autofit column widths + apply center alignment column-wide
                for col_idx, col_name in enumerate(df.columns):
                    max_len = 0
                    if not df.empty:
                        val = df[col_name].astype(str).str.len().max()
                        if pd.notna(val):
                            max_len = int(val)
                    header_len = len(str(col_name))
                    width = max(max(max_len, header_len) + 3, 12)
                    worksheet.set_column(col_idx, col_idx, width, align_fmt)

                # Apply border ONLY to the actual data rows (not the whole
                # column down to row 1,048,576) via conditional_format, which
                # lets us scope the border to exactly rows 1..len(df) in one
                # fast call per sheet instead of looping over every cell.
                if not df.empty:
                    last_row = len(df)  # 0-indexed: row 1 through last_row
                    last_col = len(df.columns) - 1
                    worksheet.conditional_format(1, 0, last_row, last_col, {
                        'type': 'formula',
                        'criteria': '=TRUE()',
                        'format': border_fmt
                    })
        return excel_buffer.getvalue()
    except Exception:
        # Fallback to openpyxl with fast formatting if xlsxwriter is missing
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            for sheet_name, df in sheet_dict.items():
                if df is None:
                    df = pd.DataFrame()
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                if sheet_name in writer.sheets:
                    format_data_sheet(writer.sheets[sheet_name], df)
        return excel_buffer.getvalue()
    
    return wb
